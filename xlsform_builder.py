import argparse
import os
from openpyxl import load_workbook, Workbook


def load_block(path):
    """Load a block workbook and return survey, choices, settings sheets as lists of rows."""
    wb = load_workbook(path)
    sheets = {}
    for name in ['survey', 'choices', 'settings']:
        if name in wb.sheetnames:
            ws = wb[name]
            sheets[name] = [list(cell.value for cell in row) for row in ws.iter_rows()]
        else:
            sheets[name] = []
    return sheets


def append_rows(ws, rows):
    for row in rows:
        ws.append(row)


def main():
    parser = argparse.ArgumentParser(description="Combine XLSForm blocks")
    parser.add_argument('blocks', nargs='+', help='Block XLSX files or names in forms directory')
    parser.add_argument('-o', '--output', required=True, help='Output XLSX file path')
    parser.add_argument('-d', '--directory', default='forms', help='Directory containing blocks')
    args = parser.parse_args()

    output_wb = Workbook()
    out_survey = output_wb.active
    out_survey.title = 'survey'
    out_choices = output_wb.create_sheet('choices')
    out_settings = output_wb.create_sheet('settings')

    first = True
    for block in args.blocks:
        if not block.endswith('.xlsx'):
            block += '.xlsx'
        path = block if os.path.exists(block) else os.path.join(args.directory, block)
        if not os.path.exists(path):
            raise FileNotFoundError(f'Cannot find block: {path}')
        sheets = load_block(path)
        if first:
            append_rows(out_survey, sheets['survey'])
            append_rows(out_choices, sheets['choices'])
            append_rows(out_settings, sheets['settings'])
            first = False
        else:
            # skip headers for subsequent blocks
            append_rows(out_survey, sheets['survey'][1:])
            append_rows(out_choices, sheets['choices'][1:])
            # ignore settings

    output_wb.save(args.output)


if __name__ == '__main__':
    main()
